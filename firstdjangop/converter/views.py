from django.shortcuts import render
from django.http import HttpResponse, FileResponse
from openpyxl import load_workbook
from vobject import vCard

def home(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        prefix_option = request.POST.get('prefix_option')
        prefix_text = request.POST.get('prefix_text')
        postfix_option = request.POST.get('postfix_option')
        postfix_text = request.POST.get('postfix_text')
        if prefix_option == 'yes' and prefix_text:
            prefix = prefix_text.strip()
        else:
            prefix = ''  # Default empty prefix if not provided
        if postfix_option == 'yes' and postfix_text:
            postfix = postfix_text.strip()
        else:
            postfix = ''    

        if excel_file.name.endswith('.xlsx') or excel_file.name.endswith('.xls'):
            workbook = load_workbook(excel_file)
            sheet = workbook.active

            vcf_content = ''  # Initialize vCard content
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming headers in row 1
                name = str(row[0]) if row[0] else 'No Name'  # Assuming names in the first column
                contact = str(row[1]) if row[1] else ''  # Assuming contacts in the second column

                # Create vCard
                card = vCard()
                card.add('fn').value = f"{prefix} {name} {postfix}" if prefix else name
                card.add('tel').value = contact

                # Append vCard content
                vcf_content += card.serialize()

            # Save the VCF content to a temporary file
            import os

            # Get the project directory path
            BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            vcf_files_dir = os.path.join(BASE_DIR, 'vcf_files')

            if not os.path.exists(vcf_files_dir):
                os.makedirs(vcf_files_dir)
            # Save the VCF content to a temporary file within the project directory
            temp_file_path = os.path.join(BASE_DIR, 'vcf_files', 'generated.vcf')

            with open(temp_file_path, 'w', encoding='utf-8') as vcf_file:
                vcf_file.write(vcf_content)

            # Serve the file as a download
            response = FileResponse(open(temp_file_path, 'rb'), content_type='text/vcard')
            response['Content-Disposition'] = 'attachment; filename="generated.vcf"'
            return response
        else:
            return HttpResponse("Invalid file format. Please upload an Excel file.")
    return render(request, 'home.html')
